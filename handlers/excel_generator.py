import openpyxl
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText
from copy import copy
from config.excelConfig import ExcelConfig as Config
from handlers.helper import Helper

class ExcelGenerator:
    def __init__(self, template='.\\templates\\edu_plan_template.xlsx'):
        self.template = template

    def generate(self, file_path, values):
        workbook = openpyxl.load_workbook(self.template)
        worksheet = workbook["графік"]

        study_period = values['study_period']
        rows_added = (study_period['years'] + 1 if study_period['months'] > 0 else 0) - 1
        total_rows = 1 + rows_added

        if (rows_added > 0):
            self.__add_rows_to_table(worksheet, rows_added)
            self.__normalize_worksheet(worksheet, rows_added)

        self.__set_input_values_to_worksheet(worksheet, values)
        self.__set_formulas_to_table(worksheet, total_rows)

        workbook.save(file_path)

    def __set_formulas_to_table(self, worksheet, total_rows):
        for i, row in enumerate(range(Config.TABLE_FIRST_ROW_IDX, Config.TABLE_FIRST_ROW_IDX + total_rows)):
            worksheet.cell(row, 2).value = Config.COURSE_IDXS[i]
            for j, column in enumerate(range(Config.FIRST_FORMULAS_COLUMN, Config.LAST_FORMULAS_COLUMN)):
                val = Config.ROW_FORMULAS[j].format(rowIdx = row)
                worksheet.cell(row, column).value = val

        summary_row_idx = Config.TABLE_FIRST_ROW_IDX + total_rows
        for i, column in enumerate(range(Config.FIRST_FORMULAS_COLUMN, Config.LAST_FORMULAS_COLUMN)):
            val = Config.SUMMARY_FORMULAS[i].format(
                firsRow = Config.TABLE_FIRST_ROW_IDX, 
                lastRow = Config.TABLE_FIRST_ROW_IDX + total_rows - 1)
            worksheet.cell(summary_row_idx, column).value = val

    def __add_rows_to_table(self, worksheet, rows_count):
        merged_cells_range = worksheet.merged_cells.ranges
        for merged_cell in merged_cells_range:
            _, min_row, _, _ = openpyxl.utils.range_boundaries(str(merged_cell))
            if min_row >= Config.TABLE_FIRST_ROW_IDX:
                merged_cell.shift(0, rows_count)

        worksheet.insert_rows(Config.TABLE_FIRST_ROW_IDX, amount=rows_count)

        for i in range(rows_count):
            self.__copy_row_format(
                worksheet, 
                Config.TABLE_FIRST_ROW_IDX + rows_count, 
                Config.TABLE_FIRST_ROW_IDX + i)
    
    def __normalize_worksheet(self, worksheet, rows_added_count):
        for rowIdx in range(Config.TABLE_FIRST_ROW_IDX, Config.LAST_ROW_IDX + rows_added_count):
            worksheet.row_dimensions[rowIdx].height = None
            if (rowIdx == Config.SIGNS_ROW_IDX_1 + rows_added_count or 
                rowIdx == Config.SIGNS_ROW_IDX_2 + rows_added_count):
                worksheet.row_dimensions[rowIdx].height = 15.75
                continue

            if rowIdx == Config.SIGNS_ROW_IDX_BTW + rows_added_count:
                worksheet.row_dimensions[rowIdx].height = 9.75

            if rowIdx == Config.PRACT_HEADERS_ROW_IDX + rows_added_count:
                worksheet.row_dimensions[rowIdx].height = 30
                continue

    def __set_input_values_to_worksheet(self, worksheet, values):
        start_edu_year = int(values['start_edu_date'])
        end_edu_year = start_edu_year + values['study_period']['years'] + 1 if values['study_period']['months'] > 0 else 0
        
        worksheet["X9"] = CellRichText(
            TextBlock(
                InlineFont(rFont="Times New Roman", sz=12), 
                f'на  {start_edu_year} - {end_edu_year}  навчальні роки'
            ),
        )

        worksheet["J10"] = CellRichText(
            TextBlock(
                InlineFont(rFont="Times New Roman", sz=12), 
                f'''за освітньо-професійною програмою "{values['study_program']}"'''
            ),
        )

        worksheet["B11"] = CellRichText(
            TextBlock(
                InlineFont(rFont="Times New Roman", sz=10, b=True), 'підготовки   '
            ),
            TextBlock(
                InlineFont(rFont="Times New Roman", sz=10, u="single"), 
                values['study_level']['name_genitive']
            ),
        )

        worksheet["B12"] = CellRichText(
            TextBlock(
                InlineFont(rFont="Times New Roman", sz=10, b=True), 'галузь знань  '
            ),
            TextBlock(
                InlineFont(rFont="Times New Roman", sz=10, u="single"), 
                f"{values['discipline']['code']} {values['discipline']['name']}"
            ),
        )

        worksheet["B13"] = CellRichText(
            TextBlock(
                InlineFont(rFont="Times New Roman", sz=10, b=True), 'форма навчання    '
            ),
            TextBlock(
                InlineFont(rFont="Times New Roman", sz=10, u="single"), values['study_form']
            )
        )

        worksheet["B15"] = CellRichText(
            TextBlock(
                InlineFont(rFont="Times New Roman", sz=10, b=True), 
                'спеціальність   '
            ),
            TextBlock(
                InlineFont(rFont="Times New Roman", sz=10, u="single"), 
                f"{values['speciality']['code']} {values['speciality']['name']}"
            )
        )

        worksheet["AQ11"] = CellRichText(
            TextBlock(
                InlineFont(rFont="Times New Roman", sz=10, b=True), 
                'освітня кваліфікація '
            ),
            TextBlock(
                InlineFont(rFont="Times New Roman", sz=10, u="single"), 
                f'''"{values['study_level']['name']} з {values['speciality']['name_genitive'].lower()}"'''
            )
        )

        worksheet["AQ13"] = CellRichText(
            TextBlock(
                InlineFont(rFont="Times New Roman", sz=10, b=True), 
                'строк навчання   '
            ),
            TextBlock(
                InlineFont(rFont="Times New Roman", sz=10, u="single"), 
                f"{Helper.year_declension(values['study_period']['years'])} {Helper.month_declension(values['study_period']['months'])}"
            )
        )

        worksheet["AQ14"] = CellRichText(
            TextBlock(
                InlineFont(rFont="Times New Roman", sz=10, b=True), 'на основі   '
            ),
            TextBlock(
                InlineFont(rFont="Times New Roman", sz=7.5), 
                'повної загальної середньої освіти (3 рівень НРК) або вищого рівня'
            )
        )

    def __copy_row_format(self, ws, source_row, new_row):
        for col in range(1, ws.max_column + 1):
            source_cell = ws.cell(row=source_row, column=col)
            new_cell = ws.cell(row=new_row, column=col)
            new_cell.font = copy(source_cell.font)
            new_cell.border = copy(source_cell.border)
            new_cell.fill= copy(source_cell.fill)
            new_cell.alignment = copy(source_cell.alignment)
            new_cell.number_format = copy(source_cell.number_format)
            new_cell.protection = copy(source_cell.protection)

    def __delete_row_with_merged_ranges(self, worksheet, idx):
        worksheet.delete_rows(idx)
        for mcr in worksheet.merged_cells:
            if idx < mcr.min_row:
                mcr.shift(row_shift=-1)
            elif idx <= mcr.max_row:
                mcr.shrink(bottom=1)